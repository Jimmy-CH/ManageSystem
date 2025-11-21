import io
import qrcode
import os
import logging
import pandas as pd
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font, Alignment, PatternFill
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView
from drf_spectacular.utils import extend_schema
from django.db.models import Prefetch
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView

from .filters import RoleFilter, UserFilter
from .models import User, Role, CustomPermission
from .serializers import UserSerializer, RoleSerializer, CustomPermissionSerializer, RegisterSerializer, \
    CustomTokenObtainPairSerializer
from utils import StandardResponse
from urllib.parse import urlparse


logger = logging.getLogger('ms')


class CustomPermissionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    只读权限列表，供前端选择
    """
    queryset = CustomPermission.objects.all().order_by('category', 'name')
    serializer_class = CustomPermissionSerializer
    pagination_class = None  # 权限不多，不分页

    @action(detail=False, methods=['get'])
    def grouped(self, request):
        """
        按 category 分组返回权限
        """
        perms = self.get_queryset()
        grouped = {}
        for p in perms:
            cat = p.category
            if cat not in grouped:
                grouped[cat] = []
            grouped[cat].append(CustomPermissionSerializer(p).data)
        return StandardResponse(data=grouped)

    @action(detail=False, methods=['get'])
    def all(self, request):
        """
        返回所有权限
        """
        queryset = self.get_queryset().values('id', 'name')
        return StandardResponse(data=list(queryset))


class RoleViewSet(viewsets.ModelViewSet):

    queryset = Role.objects.prefetch_related(
        Prefetch(
            'permissions',
            queryset=CustomPermission.objects.only('id', 'name').order_by('name')
        )
    ).all().order_by('name')
    serializer_class = RoleSerializer
    filterset_class = RoleFilter  # ← 启用过滤器
    search_fields = ['name', 'description']
    ordering_fields = ['id', 'name', 'importance', 'create_time']
    ordering = ['id']  # 默认排序

    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        导出角色数据为 Excel
        支持按当前筛选条件导出（过滤、搜索、排序）
        """
        queryset = self.filter_queryset(self.get_queryset())

        data = []
        for role in queryset:
            permission_names = ', '.join([p.name for p in role.permissions.all()])
            data.append({
                'ID': role.id,
                '角色名称': role.name,
                '描述': role.description or '',
                '状态': '启用' if role.status else '禁用',
                '重要程度': role.importance,
                '权限列表': permission_names,
                '创建时间': role.create_time.strftime('%Y-%m-%d %H:%M:%S') if role.create_time else '',
                '更新时间': role.update_time.strftime('%Y-%m-%d %H:%M:%S') if role.update_time else '',
            })
        df = pd.DataFrame(data)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="角色列表.xlsx"'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='角色数据')

        return response

    @action(detail=False, methods=['get'])
    def all(self, request):
        """
        返回所有权限
        """
        queryset = self.get_queryset().values('id', 'name')
        return StandardResponse(data=list(queryset))


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.prefetch_related('roles__permissions').all().order_by('-create_time')
    serializer_class = UserSerializer
    filterset_class = UserFilter
    search_fields = ['username', 'phone']
    ordering_fields = ['id', 'username', 'importance', 'create_time']
    ordering = ['id']  # 默认排序

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        return super().update(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='upload-avatar')
    def upload_avatar(self, request, pk=None):
        """
        上传头像：POST /users/{id}/upload-avatar/
        请求体：multipart/form-data，字段名为 avatar
        """
        user = self.get_object()
        avatar = request.FILES.get('avatar')

        if not avatar:
            return Response(
                {'error': '请上传头像文件'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 可选：验证文件类型（如只允许图片）
        if not avatar.content_type.startswith('image/'):
            return Response(
                {'error': '仅支持图片格式'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 删除旧头像（如果存在且不是默认头像）
        if user.avatar:
            old_path = user.avatar.path
            if os.path.isfile(old_path):
                default_storage.delete(old_path)

        # 保存新头像
        user.avatar.save(avatar.name, ContentFile(avatar.read()), save=True)

        return Response(
            {'avatar': user.avatar.url},
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['delete'], url_path='delete-avatar')
    def delete_avatar(self, request, pk=None):
        """
        删除头像：DELETE /users/{id}/delete-avatar/
        """
        user = self.get_object()

        if user.avatar:
            # 删除文件系统中的文件
            if default_storage.exists(user.avatar.name):
                default_storage.delete(user.avatar.name)
            # 清空模型字段
            user.avatar.delete(save=True)

        return Response(
            {'message': '头像已删除'},
            status=status.HTTP_204_NO_CONTENT
        )

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """
        使用 pandas + openpyxl 导出用户数据为 Excel
        支持过滤参数：?department=技术部&status=true
        """
        queryset = self.filter_queryset(self.get_queryset())

        data = []
        for user in queryset:
            role_names = ", ".join(user.roles.values_list('name', flat=True)) if user.roles.exists() else ""
            permissions = ", ".join(user.all_permissions) if hasattr(user, 'all_permissions') else ""

            data.append({
                "ID": user.id,
                "用户名": user.username,
                "邮箱": user.email or "",
                "手机号": user.phone or "",
                "部门": user.department or "",
                "职位": user.position or "",
                "状态": "启用" if user.status else "禁用",
                "重要程度": user.importance,
                "角色列表": role_names,
                "权限列表": permissions,
                "创建时间": user.create_time.strftime("%Y-%m-%d %H:%M:%S") if user.create_time else "",
                "更新时间": user.update_time.strftime("%Y-%m-%d %H:%M:%S") if user.update_time else "",
            })

        df = pd.DataFrame(data)

        if df.empty:
            return Response({"error": "无符合条件的数据"}, status=404)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"用户数据_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='用户数据', index=False, startrow=0, startcol=0)

            workbook = writer.book
            worksheet = writer.sheets['用户数据']

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for col_num, column_title in enumerate(df.columns, 1):
                column_letter = get_column_letter(col_num)
                max_length = max(
                    len(str(column_title)),
                    df[column_title].astype(str).map(len).max() if not df[column_title].empty else 0,
                    10  # 最小宽度
                )
                adjusted_width = min(max_length + 2, 50)  # 最大宽度 50
                worksheet.column_dimensions[column_letter].width = adjusted_width

        return response


class RegisterView(APIView):
    """
    用户注册
    """
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = RegisterSerializer(data=request.data)
        # drf框架捕获异常
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response({
            "message": "注册成功",
            "user": UserSerializer(user).data
        }, status=status.HTTP_201_CREATED)


class LoginView(TokenObtainPairView):
    """
    用户登录（返回 JWT Token + 用户信息）
    """
    serializer_class = CustomTokenObtainPairSerializer
    permission_classes = [AllowAny]


class LogoutView(APIView):
    """
    用户登出（将 refresh token 加入黑名单）
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data["refresh"]
            token = RefreshToken(refresh_token)
            token.blacklist()

            return StandardResponse(code=205, message="登出成功")
        except Exception as e:
            logger.error(f"登出失败，error: {e}")
            return StandardResponse(code=508, message="无效的 token")


class MeView(APIView):
    """
    获取当前用户信息
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = UserSerializer(request.user)
        return StandardResponse(data=serializer.data)


class CustomPermissionCreateView(APIView):
    permission_classes = [IsAdminUser]  # 仅管理员可访问

    @extend_schema(
        request=CustomPermissionSerializer,
        responses=CustomPermissionSerializer
    )
    def post(self, request):
        # 只有管理员才能创建权限
        serializer = CustomPermissionSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)


class QRCodeAPIView(APIView):
    permission_classes = [AllowAny]
    authentication_classes = []  # 如果不需要认证

    def get(self, request, *args, **kwargs):
        data = request.query_params.get('data')
        if not data:
            return StandardResponse(
                message="Missing 'data' parameter",
                status=status.HTTP_400_BAD_REQUEST
            )

        if not self.is_valid_url(data):
            return StandardResponse(
                message="Only valid HTTP/HTTPS URLs are allowed",
                status=status.HTTP_400_BAD_REQUEST
            )

        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")

        buffer = io.BytesIO()
        img.save(buffer, format="PNG")
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type="image/png")
        response["Content-Disposition"] = 'inline; filename="qrcode.png"'
        return response

    def is_valid_url(self, url):
        try:
            result = urlparse(url)
            return all([result.scheme in ('http', 'https'), result.netloc])
        except Exception as e:
            logger.error(f'error: {e}')
            return False

