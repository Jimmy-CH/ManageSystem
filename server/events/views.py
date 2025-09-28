from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter

from .constants import PRIORITY_CHOICES
from .models import Category, SLAStandard, Incident, Fault
from .serializers import (
    CategorySerializer, SLAStandardSerializer, IncidentListSerializer,
    IncidentDetailSerializer, IncidentCreateUpdateSerializer,
    FaultSerializer, CategoryTreeSerializer
)
from .filters import IncidentFilter, FaultFilter, CategoryFilter
from django.utils import timezone
from .permissions import IncidentPermission
from django.db.models import Count, Avg, F, ExpressionWrapper, DurationField, Q, Prefetch
from datetime import timedelta


class CategoryViewSet(viewsets.ModelViewSet):
    permission_classes = [IncidentPermission]  # 添加权限控制
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = CategoryFilter
    ordering_fields = ['order', 'id', 'level']
    ordering = ['order', 'id']

    @action(detail=False, methods=['get'], url_path='tree')
    def tree_view(self, request):
        categories = Category.objects.select_related('parent').prefetch_related(
            Prefetch('children', queryset=Category.objects.order_by('order', 'id'))
        ).order_by('order', 'id')

        cat_map = {cat.id: cat for cat in categories}
        for cat in categories:
            cat._children_list = []
        for cat in categories:
            if cat.parent_id in cat_map:
                cat_map[cat.parent_id]._children_list.append(cat)

        roots = [cat for cat in categories if cat.parent_id is None]
        serializer = CategoryTreeSerializer(roots, many=True)
        return Response(serializer.data)


class SLAStandardViewSet(viewsets.ModelViewSet):
    permission_classes = [IncidentPermission]  # 添加权限控制
    queryset = SLAStandard.objects.all()
    serializer_class = SLAStandardSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['priority']
    ordering_fields = ['priority']
    ordering = ['priority']


class IncidentViewSet(viewsets.ModelViewSet):
    permission_classes = [IncidentPermission]  # 添加权限控制
    queryset = Incident.objects.select_related(
        'category', 'reporter', 'assignee', 'sla'
    ).prefetch_related('faults')
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = IncidentFilter
    search_fields = ['title', 'description']
    ordering_fields = ['created_at', 'occurred_at', 'priority', 'status']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return IncidentListSerializer
        elif self.action == 'retrieve':
            return IncidentDetailSerializer
        return IncidentCreateUpdateSerializer

    @action(detail=True, methods=['get'], url_path='faults')
    def faults(self, request, pk=None):
        """
        获取指定事件下的所有故障列表
        """
        incident = self.get_object()  # 自动处理权限和404
        faults = incident.faults.all()
        serializer = FaultSerializer(faults, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='respond')
    def mark_responded(self, request, pk=None):
        """快速标记为已响应"""
        incident = self.get_object()
        if not incident.responded_at:
            incident.responded_at = timezone.now()
            incident.save(update_fields=['responded_at'])
        serializer = self.get_serializer(incident)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='resolve')
    def mark_resolved(self, request, pk=None):
        """快速标记为已解决"""
        incident = self.get_object()
        if incident.status not in [2, 3]:
            incident.status = 2
            incident.resolved_at = timezone.now()
            incident.save(update_fields=['status', 'resolved_at'])
        serializer = self.get_serializer(incident)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='statistics/general')
    def general_statistics(self, request):
        """总体统计"""
        total = self.queryset.count()
        unresolved = self.queryset.filter(status__in=[0, 1]).count()
        overdue = self.queryset.filter(
            Q(responded_at__isnull=True, sla__response_time__isnull=False) |
            Q(resolved_at__isnull=True, sla__resolve_time__isnull=False, status__in=[0, 1])
        ).count()
        avg_resolve_time = self.queryset.filter(
            resolved_at__isnull=False
        ).aggregate(
            avg=Avg(ExpressionWrapper(F('resolved_at') - F('occurred_at'), output_field=DurationField()))
        )['avg']
        print(avg_resolve_time)

        return Response({
            "total_incidents": total,
            "unresolved": unresolved,
            "overdue": overdue,
            "average_resolve_time_hours": avg_resolve_time.total_seconds() / 3600 if avg_resolve_time else 0,
        })

    @action(detail=False, methods=['get'], url_path='statistics/by-category')
    def stats_by_category(self, request):
        """按分类统计故障数"""
        data = self.queryset.values(
            'category__id', 'category__name'
        ).annotate(
            count=Count('id')
        ).order_by('-count')
        return Response(data)

    @action(detail=False, methods=['get'], url_path='statistics/by-priority')
    def stats_by_priority(self, request):
        """按优先级统计"""
        data = self.queryset.values('priority').annotate(
            count=Count('id')
        ).order_by('priority')
        for item in data:
            item['priority_display'] = dict(PRIORITY_CHOICES).get(item['priority'])
        return Response(data)

    @action(detail=False, methods=['get'], url_path='statistics/trend')
    def incident_trend(self, request):
        """
        返回最近30天每天发生的事件数量（按用户本地时区统计）
        字段：occurred_at（DateTimeField，可能为 None）
        """
        # 1. 确定本地时区的日期范围
        now_local = timezone.localtime(timezone.now())
        end_date = now_local.date()  # 今天（本地）
        start_date = end_date - timedelta(days=29)  # 包含今天共30天

        # 2. 计算 UTC 时间范围（用于高效数据库过滤）
        #    避免拉取全表数据
        utc_start = timezone.make_aware(
            timezone.datetime.combine(start_date, timezone.datetime.min.time()),
            timezone.get_default_timezone()
        )
        utc_start = utc_start.astimezone(timezone.utc)

        utc_end = timezone.make_aware(
            timezone.datetime.combine(end_date, timezone.datetime.max.time()),
            timezone.get_default_timezone()
        )
        utc_end = utc_end.astimezone(timezone.utc)

        # 3. 从数据库拉取可能相关的记录（带时区的 occurred_at）
        raw_datetimes = self.queryset.filter(
            occurred_at__isnull=False,
            occurred_at__gte=utc_start,
            occurred_at__lte=utc_end
        ).values_list('occurred_at', flat=True)

        # 4. 在 Python 中按本地日期分组
        count_by_date = defaultdict(int)
        local_tz = timezone.get_current_timezone()

        for dt in raw_datetimes:
            # 转为本地时间并提取日期
            local_date = dt.astimezone(local_tz).date()
            # 只统计在 [start_date, end_date] 范围内的
            if start_date <= local_date <= end_date:
                count_by_date[local_date] += 1

        # 5. 补全缺失日期（连续30天）
        result = []
        current = start_date
        while current <= end_date:
            result.append({
                "date": current.isoformat(),  # 转为字符串，前端友好
                "count": count_by_date.get(current, 0)
            })
            current += timedelta(days=1)

        return Response(result)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """导出事件列表为 Excel"""
        queryset = self.filter_queryset(self.get_queryset())[:1000]  # 防止过大

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "故障事件列表"

        headers = [
            "ID", "标题", "分类", "优先级", "状态", "来源",
            "上报人", "处理人", "发生时间", "响应时间", "解决时间",
            "是否响应超时", "是否解决超时", "SLA等级", "创建时间"
        ]
        ws.append(headers)

        for obj in queryset:
            ws.append([
                obj.id,
                obj.title,
                obj.category.name if obj.category else "",
                obj.get_priority_display(),
                obj.get_status_display(),
                obj.get_source_display(),
                obj.reporter.username if obj.reporter else "",
                obj.assignee.username if obj.assignee else "",
                obj.occurred_at.strftime("%Y-%m-%d %H:%M") if obj.occurred_at else "",
                obj.responded_at.strftime("%Y-%m-%d %H:%M") if obj.responded_at else "",
                obj.resolved_at.strftime("%Y-%m-%d %H:%M") if obj.resolved_at else "",
                "是" if obj.is_overdue_response else "否",
                "是" if obj.is_overdue_resolve else "否",
                obj.sla.level_name if obj.sla else "",
                obj.created_at.strftime("%Y-%m-%d %H:%M")
            ])

        # 自动列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        response = HttpResponse(content_type='application/vnd.openxmlformats-officed.com')
        response['Content-Disposition'] = 'attachment; filename=incidents.xlsx'
        wb.save(response)
        return response


class FaultViewSet(viewsets.ModelViewSet):
    permission_classes = [IncidentPermission]  # 添加权限控制
    queryset = Fault.objects.select_related('incident')
    serializer_class = FaultSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = FaultFilter
    ordering_fields = ['created_at', 'downtime_minutes']
    ordering = ['-created_at']





