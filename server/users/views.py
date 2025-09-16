from openpyxl.utils import get_column_letter
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView

from .filters import RoleFilter, UserFilter
from .models import User, Role, CustomPermission
from .serializers import UserSerializer, RoleSerializer, CustomPermissionSerializer, RegisterSerializer, \
    CustomTokenObtainPairSerializer
from rest_framework.views import APIView
from drf_spectacular.utils import extend_schema
from django.db.models import Prefetch
from utils import StandardResponse
import logging
import pandas as pd
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font, Alignment, PatternFill


logger = logging.getLogger(__name__)


class CustomPermissionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    只读权限列表，供前端选择
    """
    queryset = CustomPermission.objects.all().order_by('category', 'name')
    serializer_class = CustomPermissionSerializer
    pagination_class = None  # 权限不多，不分页

    @action(detail=False, methods=['get'])
    def grouped(self, request):
        """
        按 category 分组返回权限
        """
        perms = self.get_queryset()
        grouped = {}
        for p in perms:
            cat = p.category
            if cat not in grouped:
                grouped[cat] = []
            grouped[cat].append(CustomPermissionSerializer(p).data)
        return StandardResponse(data=grouped)

    @action(detail=False, methods=['get'])
    def all(self, request):
        """
        返回所有权限
        """
        queryset = self.get_queryset().values('id', 'name')
        return StandardResponse(data=list(queryset))


class RoleViewSet(viewsets.ModelViewSet):

    queryset = Role.objects.prefetch_related(
        Prefetch(
            'permissions',
            queryset=CustomPermission.objects.only('id', 'name').order_by('name')
        )
    ).all().order_by('name')
    serializer_class = RoleSerializer
    filterset_class = RoleFilter  # ← 启用过滤器
    search_fields = ['name', 'description']
    ordering_fields = ['id', 'name', 'importance', 'created_at']
    ordering = ['id']  # 默认排序

    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        导出角色数据为 Excel
        支持按当前筛选条件导出（过滤、搜索、排序）
        """
        queryset = self.filter_queryset(self.get_queryset())

        data = []
        for role in queryset:
            permission_names = ', '.join([p.name for p in role.permissions.all()])
            data.append({
                'ID': role.id,
                '角色名称': role.name,
                '描述': role.description or '',
                '状态': '启用' if role.status else '禁用',
                '重要程度': role.importance,
                '权限列表': permission_names,
                '创建时间': role.created_at.strftime('%Y-%m-%d %H:%M:%S') if role.created_at else '',
                '更新时间': role.updated_at.strftime('%Y-%m-%d %H:%M:%S') if role.updated_at else '',
            })
        df = pd.DataFrame(data)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="角色列表.xlsx"'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='角色数据')

        return response

    @action(detail=False, methods=['get'])
    def all(self, request):
        """
        返回所有权限
        """
        queryset = self.get_queryset().values('id', 'name')
        return StandardResponse(data=list(queryset))


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.prefetch_related('roles__permissions').all().order_by('-created_at')
    serializer_class = UserSerializer
    filterset_class = UserFilter
    search_fields = ['username', 'phone']
    ordering_fields = ['id', 'username', 'importance', 'created_at']
    ordering = ['id']  # 默认排序

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        return super().update(request, *args, **kwargs)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """
        使用 pandas + openpyxl 导出用户数据为 Excel
        支持过滤参数：?department=技术部&status=true
        """
        # 1. 获取过滤后的数据
        queryset = self.filter_queryset(self.get_queryset())

        # 2. 构建数据字典列表
        data = []
        for user in queryset:
            role_names = ", ".join(user.roles.values_list('name', flat=True)) if user.roles.exists() else ""
            permissions = ", ".join(user.all_permissions) if hasattr(user, 'all_permissions') else ""

            data.append({
                "ID": user.id,
                "用户名": user.username,
                "邮箱": user.email or "",
                "手机号": user.phone or "",
                "部门": user.department or "",
                "职位": user.position or "",
                "状态": "启用" if user.status else "禁用",
                "重要程度": user.importance,
                "角色列表": role_names,
                "权限列表": permissions,
                "创建时间": user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user.created_at else "",
                "更新时间": user.updated_at.strftime("%Y-%m-%d %H:%M:%S") if user.updated_at else "",
            })

        # 3. 创建 DataFrame
        df = pd.DataFrame(data)

        # 4. 如果无数据，返回空文件或提示
        if df.empty:
            return Response({"error": "无符合条件的数据"}, status=404)

        # 5. 创建 Excel 响应
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"用户数据_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'

        # 6. 用 pandas 写入 Excel（使用 openpyxl 引擎）
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='用户数据', index=False, startrow=0, startcol=0)

            # 7. 获取 workbook 和 worksheet 对象，用于样式设置
            workbook = writer.book
            worksheet = writer.sheets['用户数据']

            # 8. 设置表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # 9. 自动调整列宽
            for col_num, column_title in enumerate(df.columns, 1):
                column_letter = get_column_letter(col_num)
                max_length = max(
                    len(str(column_title)),
                    df[column_title].astype(str).map(len).max() if not df[column_title].empty else 0,
                    10  # 最小宽度
                )
                adjusted_width = min(max_length + 2, 50)  # 最大宽度 50
                worksheet.column_dimensions[column_letter].width = adjusted_width

        return response

    @action(detail=False, methods=['get'])
    def me(self, request):
        serializer = self.get_serializer(request.user)
        return StandardResponse(data=serializer.data)


class RegisterView(APIView):
    """
    用户注册
    """
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = RegisterSerializer(data=request.data)
        # drf框架捕获异常
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response({
            "message": "注册成功",
            "user": UserSerializer(user).data
        }, status=status.HTTP_201_CREATED)
        # return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LoginView(TokenObtainPairView):
    """
    用户登录（返回 JWT Token + 用户信息）
    """
    serializer_class = CustomTokenObtainPairSerializer
    permission_classes = [AllowAny]


class LogoutView(APIView):
    """
    用户登出（将 refresh token 加入黑名单）
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data["refresh"]
            token = RefreshToken(refresh_token)
            token.blacklist()

            return StandardResponse(code=205, message="登出成功")
        except Exception as e:
            logger.error(f"登出失败，error: {e}")
            return StandardResponse(code=508, message="无效的 token")


class MeView(APIView):
    """
    获取当前用户信息
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = UserSerializer(request.user)
        return StandardResponse(data=serializer.data)


class CustomPermissionCreateView(APIView):
    permission_classes = [IsAdminUser]  # 仅管理员可访问

    @extend_schema(
        request=CustomPermissionSerializer,
        responses=CustomPermissionSerializer
    )
    def post(self, request):
        # 只有管理员才能创建权限
        serializer = CustomPermissionSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)
